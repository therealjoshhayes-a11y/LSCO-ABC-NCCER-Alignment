# ingest_wecm.py
# Antikythera Pipeline — LSCO ABC NCCER Alignment
# WECM catalog ingest: parse vertical multi-row records, deduplicate by
# content unit, return structured course records ready for embedding.

import openpyxl
import pandas as pd
from pathlib import Path
from code.config import WECM_FILE, PROCESSED_DIR

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

LABEL_COURSE_LEVEL   = "Course Level:"
LABEL_DESCRIPTION    = "Course Description:"
LABEL_OUTCOMES       = "End-of-Course Outcomes:"
ACTIVE_STATUS        = "Active"
COLUMN_HEADER_CIP    = "CIP"


def is_cip_code(value) -> bool:
    """Check if a value looks like a CIP code e.g. '47.0603' or '46'."""
    if not value:
        return False
    s = str(value).strip()
    parts = s.replace(".", "")
    return parts.isdigit() and len(s) >= 2


def is_course_header_row(row) -> bool:
    """
    Detect a course header row.
    Criteria: CIP looks like a CIP code, Rubric is populated,
    Number is populated, Status is populated.
    """
    cip    = row[0]
    rubric = row[1]
    number = row[2]
    status = row[4]

    return (
        is_cip_code(cip)
        and rubric
        and number
        and status in (ACTIVE_STATUS, "Archived", "Inactive")
    )


# ─────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────

def parse_wecm(filepath: Path = WECM_FILE) -> list[dict]:
    """
    Parse WECM xlsx — one course per sheet, 5856 sheets total.
    Each sheet follows the same vertical multi-row structure.
    """
    print(f"Loading WECM from: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    courses = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current = {}

        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            if str(row[0]).strip() == COLUMN_HEADER_CIP:
                continue

            cip_val = str(row[0]).strip() if row[0] is not None else ""

            if is_course_header_row(row):
                current = {
                    "cip":         str(row[0]).strip(),
                    "rubric":      str(row[1]).strip(),
                    "number":      str(row[2]).strip(),
                    "title":       str(row[3]).strip() if row[3] else "",
                    "status":      str(row[4]).strip() if row[4] else "",
                    "sch":         row[5],
                    "min_hours":   row[6],
                    "max_hours":   row[7],
                    "level":       "",
                    "description": "",
                    "outcomes":    "",
                }
                continue

            if cip_val.startswith("Course Level:"):
                if current:
                    current["level"] = str(row[3]).strip() if row[3] else ""

            elif cip_val.startswith("Course Description:"):
                if current:
                    current["description"] = str(row[3]).strip() if row[3] else ""

            elif cip_val.startswith("End-of-Course Outcomes:"):
                if current:
                    current["outcomes"] = str(row[3]).strip() if row[3] else ""

        if current.get("rubric"):
            courses.append(current)

    wb.close()
    print(f"Raw records parsed: {len(courses)}")
    return courses


# ─────────────────────────────────────────────
# ACTIVE STATUS FILTER
# ─────────────────────────────────────────────

def filter_active(courses: list[dict]) -> list[dict]:
    """Keep only Active-status courses."""
    active = [c for c in courses if c.get("status") == ACTIVE_STATUS]
    print(f"Active records after filter: {len(active)}")
    return active


# ─────────────────────────────────────────────
# CONTENT UNIT DEDUPLICATION
# ─────────────────────────────────────────────

def get_content_key(course: dict) -> str:
    """
    Content-stable deduplication key.
    Format: RUBRIC_Level_ContentID
    Example: WLDG_1_57 for WLDG 1257 and WLDG 1457 (same content, different SCH)

    Number format: LSCD
      L = Level (first digit)
      S = SCH (second digit)
      CD = Content ID (last two digits)
    """
    rubric = course.get("rubric", "")
    number = course.get("number", "")
    if len(number) >= 4:
        level      = number[0]
        content_id = number[2:]
        return f"{rubric}_{level}_{content_id}"
    return f"{rubric}_{number}"


def deduplicate_by_content_unit(courses: list[dict]) -> list[dict]:
    """
    Group SCH variants under a single content unit record.
    Embed once per content unit. Retain all SCH variants as a list.
    Description and outcomes taken from first encountered record.
    """
    content_units = {}

    for course in courses:
        key = get_content_key(course)

        if key not in content_units:
            content_units[key] = {
                "content_key":  key,
                "cip":          course["cip"],
                "rubric":       course["rubric"],
                "number_base":  course["number"],
                "title":        course["title"],
                "status":       course["status"],
                "level":        course["level"],
                "description":  course["description"],
                "outcomes":     course["outcomes"],
                "sch_variants": [course["sch"]],
                "max_hours":    course["max_hours"],
                "min_hours":    course["min_hours"],
            }
        else:
            # Add SCH variant, keep highest max_hours
            content_units[key]["sch_variants"].append(course["sch"])
            if course["max_hours"] and (
                not content_units[key]["max_hours"]
                or course["max_hours"] > content_units[key]["max_hours"]
            ):
                content_units[key]["max_hours"] = course["max_hours"]

    deduped = list(content_units.values())
    print(f"Content units after deduplication: {len(deduped)}")
    return deduped


# ─────────────────────────────────────────────
# CLAIMS BLOCK CONSTRUCTION
# ─────────────────────────────────────────────

def build_claims_block(course: dict) -> str:
    """
    Concatenate description + outcomes into a single claims block string.
    If outcomes is empty, use description only.
    This is the text that gets embedded by SBERT.
    """
    desc     = course.get("description", "").strip()
    outcomes = course.get("outcomes", "").strip()

    if desc and outcomes:
        return f"{desc} {outcomes}"
    elif desc:
        return desc
    elif outcomes:
        return outcomes
    else:
        return ""


def add_claims_blocks(courses: list[dict]) -> list[dict]:
    """Add claims_block field to each content unit."""
    empty_count = 0
    for course in courses:
        course["claims_block"] = build_claims_block(course)
        if not course["claims_block"]:
            empty_count += 1
    print(f"Claims blocks built. Empty blocks: {empty_count}")
    return courses


# ─────────────────────────────────────────────
# CIP SCOPE FILTER
# ─────────────────────────────────────────────

def filter_by_cip(courses: list[dict], cip_scope: list[str]) -> list[dict]:
    """
    Filter content units to CIP scope for a given trace.
    cip_scope is a list of CIP prefixes e.g. ['46', '48', '15']
    or exact codes e.g. ['46.0300']
    """
    filtered = [
        c for c in courses
        if any(c.get("cip", "").startswith(prefix) for prefix in cip_scope)
    ]
    print(f"Records in CIP scope {cip_scope}: {len(filtered)}")
    return filtered


# ─────────────────────────────────────────────
# RUBRIC FILTER
# ─────────────────────────────────────────────

def filter_by_rubric(courses: list[dict], rubric_scope: list[str]) -> list[dict]:
    """Filter content units to rubric scope for a given trace."""
    if not rubric_scope:
        return courses
    filtered = [c for c in courses if c.get("rubric") in rubric_scope]
    print(f"Records in rubric scope {rubric_scope}: {len(filtered)}")
    return filtered


# ─────────────────────────────────────────────
# SAVE TO PROCESSED
# ─────────────────────────────────────────────

def save_processed(courses: list[dict], filename: str) -> Path:
    """Save processed course records to processed/ as CSV."""
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    outpath = PROCESSED_DIR / filename
    df = pd.DataFrame(courses)
    df.to_csv(outpath, index=False)
    print(f"Saved: {outpath} ({len(df)} records)")
    return outpath


# ─────────────────────────────────────────────
# MAIN INGEST PIPELINE
# ─────────────────────────────────────────────

def run_wecm_ingest() -> list[dict]:
    """
    Full WECM ingest pipeline:
    1. Parse vertical multi-row records
    2. Filter to Active status
    3. Deduplicate by content unit
    4. Build claims blocks
    5. Save full processed catalog

    Returns list of content unit dicts with claims blocks.
    """
    print("\n=== WECM INGEST ===")
    courses  = parse_wecm()
    courses  = filter_active(courses)
    courses  = deduplicate_by_content_unit(courses)
    courses  = add_claims_blocks(courses)
    save_processed(courses, "wecm_processed.csv")
    print("=== WECM INGEST COMPLETE ===\n")
    return courses


if __name__ == "__main__":
    run_wecm_ingest()