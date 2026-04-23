"""
coverage.py — Coverage Map Generator
Antikythera Pipeline | LSCO–ABC Alignment Project

Two modes:
  Mode 1 — Heuristic: top reordered queue_pos=1 pair per module per program.
            Runs against queue files. No panel decisions required.
            Output: {program}_coverage_heuristic.xlsx

  Mode 2 — Final: reads accepted decisions from allocation_ledger.csv.
            Primary placement = highest SBERT score among accepted allocations.
            Output: {program}_coverage_final.xlsx

Each workbook contains three tabs:
  1. Course View   — per course: all modules placed on it, SBERT scores
  2. Module View   — per module: placements, primary flag, gap flag
  3. Credit Language — draft Appendix 3 credit schedule block

Programs consolidated across all traces (Core + craft together).
"""

import math
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from code.config import PROCESSED_DIR, INTERIM_DIR
from code.ledger import QUEUE_DIR, LEDGER_DIR, LEDGER_FILE, TRACE_PROGRAMS

OUTPUT_DIR = INTERIM_DIR / "coverage"

# ── Program → traces mapping ──────────────────────────────────────────────────

PROGRAM_TRACES = {
    "electrical_aas":      ["t0_core", "t1_electrical"],
    "welding_aas":         ["t0_core", "t2_welding", "t3_pipefitting"],
    "instrumentation_aas": ["t0_core", "t4_instrumentation"],
    "bct_aas":             ["t0_core", "t5_carpentry"],
}

PROGRAM_LABELS = {
    "electrical_aas":      "Electrical AAS",
    "welding_aas":         "Welding Fabrication AAS",
    "instrumentation_aas": "Instrumentation AAS",
    "bct_aas":             "Building Construction Technology AAS",
}

TRACE_LABELS = {
    "t0_core":            "NCCER Core",
    "t1_electrical":      "Electrical L1-L4",
    "t2_welding":         "Welding L1-L2",
    "t3_pipefitting":     "Pipefitting L1-L4",
    "t4_instrumentation": "Instrumentation L1-L4",
    "t5_carpentry":       "Carpentry",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def content_key_to_course_id(content_key: str, rubric_lookup: dict) -> str:
    """
    Reverse content_key to display course_id where possible.
    CNBT_1_10 → CNBT 1210 (via degree plan lookup).
    Falls back to content_key if no match found.
    """
    return rubric_lookup.get(content_key, content_key.replace("_", " ", 1)
                             .replace("_", ""))


def build_content_key_lookup() -> dict:
    """
    Build lookup dict: content_key → course_id
    from degree_plans_processed.csv across all programs.
    """
    dp = pd.read_csv(PROCESSED_DIR / "degree_plans_processed.csv")
    lookup = {}
    for _, row in dp.iterrows():
        parts   = row["course_id"].strip().split()
        rubric  = parts[0]
        number  = parts[1]
        level   = number[0]
        content = number[2:]
        key     = f"{rubric}_{level}_{content}"
        lookup[key] = row["course_id"]
    return lookup


def safe_title(val) -> str:
    """Return string title or placeholder for NaN module titles."""
    if pd.isna(val) or str(val).strip() == "":
        return "(title unavailable — OCR artifact)"
    return str(val)


# ── Data builders ─────────────────────────────────────────────────────────────

def build_heuristic_placements(program: str) -> pd.DataFrame:
    """
    Mode 1: For each module × program, take queue_pos=1 pair from
    reordered K queue as the heuristic nominated placement.
    Consolidates across all traces for the program.
    """
    traces  = PROGRAM_TRACES.get(program, [])
    all_rows = []

    for trace_id in traces:
        queue_path = QUEUE_DIR / f"{trace_id}_queue.csv"
        if not queue_path.exists():
            continue

        df = pd.read_csv(queue_path)
        df = df[df["program"] == program].copy()

        if df.empty:
            continue

        # Top pair per module = queue_pos 1
        top = (df.sort_values("queue_pos")
                 .groupby("module_id")
                 .first()
                 .reset_index())

        top["trace_id"]          = trace_id
        top["trace_label"]       = TRACE_LABELS.get(trace_id, trace_id)
        top["allocation_source"] = "HEURISTIC"
        top["is_primary"]        = True
        all_rows.append(top)

    if not all_rows:
        return pd.DataFrame()

    return pd.concat(all_rows, ignore_index=True)


def build_final_placements(program: str) -> pd.DataFrame:
    """
    Mode 2: Read accepted decisions from allocation_ledger.csv.
    Primary placement = highest SBERT score per module.
    Consolidates across all traces for the program.
    """
    if not LEDGER_FILE.exists():
        raise FileNotFoundError(
            f"Allocation ledger not found: {LEDGER_FILE}\n"
            f"Run build_ledger() after panel decisions are recorded."
        )

    ledger = pd.read_csv(LEDGER_FILE)
    df     = ledger[ledger["program"] == program].copy()

    if df.empty:
        return pd.DataFrame()

    # Flag primary placement per module — highest SBERT score
    df["is_primary"] = False
    for (trace_id, module_id), grp in df.groupby(["trace_id", "module_id"]):
        if grp.empty:
            continue
        primary_idx = grp["sbert_score"].idxmax()
        df.at[primary_idx, "is_primary"] = True

    df["trace_label"] = df["trace_id"].map(TRACE_LABELS)
    return df


# ── Excel styling ─────────────────────────────────────────────────────────────

HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
PRIMARY_FILL   = PatternFill("solid", fgColor="E2EFDA")
ORPHAN_FILL    = PatternFill("solid", fgColor="FCE4D6")
ALT_FILL       = PatternFill("solid", fgColor="F5F5F5")
SECTION_FILL   = PatternFill("solid", fgColor="D6E4F7")
SECTION_FONT   = Font(bold=True, size=10)
THIN_BORDER    = Border(
    bottom=Side(style="thin", color="CCCCCC")
)


def style_header_row(ws, row_num: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell            = ws.cell(row=row_num, column=col)
        cell.fill       = HEADER_FILL
        cell.font       = HEADER_FONT
        cell.alignment  = Alignment(horizontal="center", wrap_text=True)


def autofit_columns(ws, min_width=10, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, min_width), max_width
        )


# ── Tab 1 — Course view ───────────────────────────────────────────────────────

def write_course_view(ws, placements: pd.DataFrame,
                      ck_lookup: dict, program: str):
    """
    For each LSCO course: all modules placed on it, SBERT score,
    trace, primary flag. Grouped by course.
    """
    ws.title = "Course View"

    headers = [
        "LSCO Course ID", "Course Title", "Trace", "NCCER Module ID",
        "Module Title", "SBERT Score", "TF-IDF Score", "Jaccard Score",
        "Promotion Delta", "Primary Placement", "Allocation Source"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    if placements.empty:
        ws.append(["No placements available."])
        return

    placements = placements.copy()
    placements["course_id_display"] = placements["wecm_course_id"].apply(
        lambda k: content_key_to_course_id(k, ck_lookup)
    )

    # Load course titles from degree plan
    dp = pd.read_csv(PROCESSED_DIR / "degree_plans_processed.csv")
    dp_prog = dp[dp["program"] == program][["course_id", "title"]].copy()
    course_title_map = dict(zip(dp_prog["course_id"], dp_prog["title"]))

    # Sort by course_id_display then trace then module_id
    placements = placements.sort_values(
        ["course_id_display", "trace_id", "module_id"]
    )

    row_num = 2
    prev_course = None

    for _, row in placements.iterrows():
        course_id = row["course_id_display"]

        if course_id != prev_course:
            # Section header for each course
            ws.append([
                course_id,
                course_title_map.get(course_id, ""),
                "", "", "", "", "", "", "", "", ""
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = SECTION_FILL
                ws.cell(row=row_num, column=col).font = SECTION_FONT
            row_num += 1
            prev_course = course_id

        data_row = [
            course_id,
            course_title_map.get(course_id, ""),
            row.get("trace_label", ""),
            row.get("module_id", ""),
            safe_title(row.get("module_title", "")),
            round(float(row.get("sbert_score", 0)), 4),
            round(float(row.get("tfidf_score", 0)), 4)
                if pd.notna(row.get("tfidf_score")) else "",
            round(float(row.get("jaccard_score", 0)), 4)
                if pd.notna(row.get("jaccard_score")) else "",
            int(row.get("promotion_delta", 0)),
            "YES" if row.get("is_primary") else "",
            row.get("allocation_source", ""),
        ]
        ws.append(data_row)

        # Highlight primary placements
        fill = PRIMARY_FILL if row.get("is_primary") else (
            ALT_FILL if row_num % 2 == 0 else None
        )
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = fill

        row_num += 1

    autofit_columns(ws)


# ── Tab 2 — Module view ───────────────────────────────────────────────────────

def write_module_view(ws, placements: pd.DataFrame,
                      ck_lookup: dict, program: str,
                      all_modules: pd.DataFrame):
    """
    For each NCCER module: placement(s), primary flag, gap flag.
    Grouped by trace then module.
    """
    ws.title = "Module View"

    headers = [
        "Trace", "Module ID", "Module Title", "LSCO Course ID",
        "Course Title", "SBERT Score", "Primary Placement",
        "Allocation Source", "Status"
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    # Load course titles
    dp = pd.read_csv(PROCESSED_DIR / "degree_plans_processed.csv")
    dp_prog = dp[dp["program"] == program][["course_id", "title"]].copy()
    course_title_map = dict(zip(dp_prog["course_id"], dp_prog["title"]))

    row_num = 2

    for _, mod_row in all_modules.sort_values(
            ["trace_id", "module_id"]).iterrows():

        trace_id  = mod_row["trace_id"]
        module_id = mod_row["module_id"]

        mod_placements = pd.DataFrame()
        if not placements.empty:
            mod_placements = placements[
                (placements["trace_id"]  == trace_id) &
                (placements["module_id"] == module_id)
            ]

        if mod_placements.empty:
            # Gap — no placement
            ws.append([
                TRACE_LABELS.get(trace_id, trace_id),
                module_id,
                safe_title(mod_row.get("module_title", "")),
                "", "", "", "", "",
                "UNALLOCATED"
            ])
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = ORPHAN_FILL
            row_num += 1
        else:
            for _, p in mod_placements.iterrows():
                course_id = content_key_to_course_id(
                    p["wecm_course_id"], ck_lookup
                )
                ws.append([
                    TRACE_LABELS.get(trace_id, trace_id),
                    module_id,
                    safe_title(p.get("module_title", "")),
                    course_id,
                    course_title_map.get(course_id, ""),
                    round(float(p.get("sbert_score", 0)), 4),
                    "YES" if p.get("is_primary") else "",
                    p.get("allocation_source", ""),
                    "PLACED"
                ])
                fill = PRIMARY_FILL if p.get("is_primary") else (
                    ALT_FILL if row_num % 2 == 0 else None
                )
                if fill:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = fill
                row_num += 1

    autofit_columns(ws)


# ── Tab 3 — Draft credit language ────────────────────────────────────────────

def write_credit_language(ws, placements: pd.DataFrame,
                          ck_lookup: dict, program: str,
                          mode: str):
    """
    Draft Appendix 3 credit schedule language block.
    Groups placements by trace, lists primary course per module,
    then consolidates to course-level credit statement.
    """
    ws.title = "Credit Language"

    prog_label = PROGRAM_LABELS.get(program, program)
    mode_label = "Heuristic Nomination (Pre-SME)" if mode == "heuristic" \
                 else "Final Panel Decisions"

    ws.append([f"DRAFT CPL Credit Schedule — {prog_label}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.append([f"Source: {mode_label}"])
    ws.cell(row=2, column=1).font = Font(italic=True, size=10)
    ws.append([""])
    ws.append(["This draft is generated from Antikythera pipeline output. "
               "Final credit language requires analyst and SME committee review "
               "before submission to program directors and THECB notification."])
    ws.cell(row=4, column=1).font = Font(italic=True, size=9)
    ws.append([""])

    row_num = 6

    if placements.empty:
        ws.append(["No placements available for credit schedule."])
        return

    # Group by trace — produce one block per credential
    for trace_id in PROGRAM_TRACES.get(program, []):
        trace_placements = placements[
            placements["trace_id"] == trace_id
        ].copy() if not placements.empty else pd.DataFrame()

        trace_label = TRACE_LABELS.get(trace_id, trace_id)

        ws.cell(row=row_num, column=1).value = f"Credential: {trace_label}"
        ws.cell(row=row_num, column=1).font  = Font(bold=True, size=11)
        ws.cell(row=row_num, column=1).fill  = SECTION_FILL
        row_num += 1

        if trace_placements.empty:
            ws.cell(row=row_num, column=1).value = "  No placements recorded."
            row_num += 2
            continue

        # Get unique courses carrying placements from this trace
        trace_placements["course_id_display"] = \
            trace_placements["wecm_course_id"].apply(
                lambda k: content_key_to_course_id(k, ck_lookup)
            )

        # Load course titles
        dp = pd.read_csv(PROCESSED_DIR / "degree_plans_processed.csv")
        dp_prog  = dp[dp["program"] == program][["course_id", "title", "credits"]]
        ct_map   = dict(zip(dp_prog["course_id"], dp_prog["title"]))
        sch_map  = dict(zip(dp_prog["course_id"], dp_prog["credits"]))

        unique_courses = sorted(
            trace_placements["course_id_display"].unique()
        )

        total_sch = sum(
            sch_map.get(c, 0) for c in unique_courses
            if pd.notna(sch_map.get(c, 0))
        )

        # Module count
        n_modules   = trace_placements["module_id"].nunique()
        n_placed    = trace_placements[
            trace_placements["course_id_display"].notna()
        ]["module_id"].nunique()

        ws.cell(row=row_num, column=1).value = \
            f"  Modules covered: {n_placed} of {n_modules}"
        row_num += 1

        ws.cell(row=row_num, column=1).value = \
            f"  Estimated CPL credit hours: {total_sch} SCH"
        row_num += 1

        ws.cell(row=row_num, column=1).value = "  Mapped to the following courses:"
        row_num += 1

        for course_id in unique_courses:
            title = ct_map.get(course_id, "")
            sch   = sch_map.get(course_id, "")
            mods  = trace_placements[
                trace_placements["course_id_display"] == course_id
            ]["module_id"].tolist()
            mod_list = ", ".join(str(m) for m in sorted(set(mods)))

            ws.cell(row=row_num, column=1).value = \
                f"    {course_id} — {title} ({sch} SCH)"
            ws.cell(row=row_num, column=2).value = \
                f"Modules: {mod_list}"
            row_num += 1

        row_num += 1  # blank line between credentials

    # Consolidated summary line
    ws.cell(row=row_num, column=1).value = \
        "─" * 80
    row_num += 1

    all_courses = []
    if not placements.empty:
        placements["course_id_display"] = placements["wecm_course_id"].apply(
            lambda k: content_key_to_course_id(k, ck_lookup)
        )
        all_courses = sorted(placements["course_id_display"].unique().tolist())

    dp      = pd.read_csv(PROCESSED_DIR / "degree_plans_processed.csv")
    dp_prog = dp[dp["program"] == program][["course_id", "credits"]]
    sch_map = dict(zip(dp_prog["course_id"], dp_prog["credits"]))

    total_sch = sum(
        sch_map.get(c, 0) for c in all_courses
        if pd.notna(sch_map.get(c, 0))
    )

    credential_names = " + ".join(
        TRACE_LABELS.get(t, t)
        for t in PROGRAM_TRACES.get(program, [])
    )
    course_list = ", ".join(all_courses)

    ws.cell(row=row_num, column=1).value = \
        f"DRAFT: {credential_names} → {total_sch} SCH credit across: {course_list}"
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=10)
    row_num += 1

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 40


# ── Main workbook builder ─────────────────────────────────────────────────────

def build_coverage_map(program: str, mode: str = "heuristic"):
    """
    Build coverage map workbook for one program.
    mode: 'heuristic' or 'final'
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ck_lookup = build_content_key_lookup()

    print(f"\n  Building {mode} coverage map: {program}")

    # Get placements
    if mode == "heuristic":
        placements = build_heuristic_placements(program)
    else:
        placements = build_final_placements(program)

    # Get all modules across all traces for this program
    all_modules_list = []
    for trace_id in PROGRAM_TRACES.get(program, []):
        queue_path = QUEUE_DIR / f"{trace_id}_queue.csv"
        if not queue_path.exists():
            continue
        df = pd.read_csv(queue_path)
        df = df[df["program"] == program][
            ["trace_id", "module_id", "module_title"]
        ].drop_duplicates()
        all_modules_list.append(df)

    all_modules = pd.concat(
        all_modules_list, ignore_index=True
    ).drop_duplicates() if all_modules_list else pd.DataFrame()

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    ws1 = wb.create_sheet("Course View")
    write_course_view(ws1, placements, ck_lookup, program)

    ws2 = wb.create_sheet("Module View")
    write_module_view(ws2, placements, ck_lookup, program, all_modules)

    ws3 = wb.create_sheet("Credit Language")
    write_credit_language(ws3, placements, ck_lookup, program, mode)

    filename = f"{program}_coverage_{mode}.xlsx"
    out_path = OUTPUT_DIR / filename
    wb.save(out_path)
    print(f"  Saved: {out_path.name}")
    return out_path


def build_all_coverage_maps(mode: str = "heuristic"):
    """Build coverage map workbooks for all programs."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"\n=== COVERAGE MAP BUILD — {mode.upper()} ===")

    output_files = []
    for program in PROGRAM_TRACES:
        path = build_coverage_map(program, mode)
        output_files.append(path)

    print(f"\n=== COVERAGE MAP COMPLETE — {len(output_files)} workbooks ===")
    print(f"Output: {OUTPUT_DIR}")
    return output_files


if __name__ == "__main__":
    build_all_coverage_maps(mode="heuristic")